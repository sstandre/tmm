# -*- coding: utf-8 -*-
"""
Created on Thu Mar 26 17:45:17 2020

@author: simon

Utilities to calculate RAT of optical stacks with tmm_vec
"""

import numpy as np
import os
import pickle
import tmm
from functools import partial
from itertools import product
from scipy.interpolate import interp1d

try:
    from openpyxl import load_workbook
    import matplotlib.pyplot as plt
    import matplotlib.cm as colors
except ImportError:
    pass

DEG = 180/np.pi

def cauchy_fn(A,B,C):
    """Modelo espectroscopico de Cauchy para n(λ).
    """
    def fn(lams):
        lams = lams*1.0
        return A+B*np.power(10,4)/np.power(lams,2)+C*np.power(10,9)/np.power(lams,4)
    return fn


def constant_fn(val):
    """Función constante.
    """
    return partial(np.full_like, fill_value=val)


def brugg_fn(n_a, n_b, f_b):
    """Modelo de campo medio de Bruggeman. n efectivo correspondiente a una mezcla con
    fracción f_b de la especie b (indices n_b) y (1-f_b) de la especie a (indices n_a).
    """
    def fn(lams):
        n1 = n_a(lams)
        n2 = n_b(lams)
        return n_eff(n1, n2, f_b)
    return fn

def load_fn(filename, skiprows=1):
    """Carga un archivo de texto de dos columnas (x, y) y devuelve una funcion y(x) que 
    interpola los valores del archivo.
    """
    wv, I = np.loadtxt(filename, skiprows=skiprows, unpack=True)
    return interp1d(wv, I,fill_value=(I[0],I[-1]), bounds_error=False)


def n_eff(n1,n2,f):
    """
    Modelo de bruggeman. f es la fraccion de la especie 2.
    """
    if f == 0: return n1
    if f == 1: return n2
    e1 = n1**2
    e2 = n2**2
    #f = 1-f2
    omega = (1-f)*(e2-2*e1)+f*(e1-2*e2)
    return (np.sqrt(np.sqrt(omega**2 + 8*e1*e2)-omega))/2


def load_interp(filename, comments=';', skiprows=0, unit='nm'):
    
    """Carga un archivo de texto de dos columnas (x, n, k) y devuelve dos funciones 
    n(x), k(x), que interpolan los valores del archivo.
    """
    scale = {'nm':1.0, 'um':1000.0}
    
    wv, n, k = np.loadtxt(filename, comments=comments, 
                      skiprows=skiprows, unpack=True)
    
    wv *= scale[unit]
    
    n_fn = interp1d(wv, n,fill_value=(n[0],n[-1]), bounds_error=False)
    k_fn = interp1d(wv, k,fill_value=(k[0],k[-1]), bounds_error=False)
    
    return n_fn, k_fn


def load_psi_delta_from_spe(filename, offset):
    """Cargar datos elipsometricos de archivos .spe del elipsometro Horiba"""
    A = np.loadtxt(filename,skiprows=offset+62, usecols=(0,1,2), encoding="latin-1")
    L= A[:-2,0]
    P = A[:-2,:2]
    D = np.c_[L, A[:-2,2]]
    return P, D


def stack2tmm(stack, materials, lams, add_inf=False):
    """Transforma una estructura en forma de  'stack' en tres listas que sirven como
    input para las funciones de la librería tmm.
    """
    n_list = []
    d_list = []
    c_list = []
    if add_inf:
        n_list.append([1.]*len(lams))
        d_list.append(np.inf)
        c_list.append('i')
        
    for layer in stack:
        thickness, name, coh = layer
        n, k = materials[name]

        d_list.append(thickness)
        n_list.append(n(lams) + k(lams)*1.0j)
        c_list.append(coh)
        
    if add_inf:  
        n_list.append([1.]*len(lams))
        d_list.append(np.inf)
        c_list.append('i')
    
    return n_list, d_list, c_list

def calculate_RT(stack, materials, lams, pol='s', th_0=0, thicks=None):
    """Calcular R y T para un stack, para todas las combinaciones de espesores."""

    n_list, d_list, c_list = stack2tmm(stack, materials, lams, add_inf=False)
    n_lams = len(lams)
    inc = 'i' in c_list[1:-1] #check if  any of the finite layers is incoherent
    if inc:
        def f(pol, n_list, d_list, th_0, lams, c_list):
            return tmm.inc_tmm(pol, n_list, d_list, c_list, th_0, lams)
    else:
        def f(pol, n_list, d_list, th_0, lams, c_list=None):
            return tmm.coh_tmm(pol, n_list, d_list, th_0, lams)
            
    if thicks is None:
        #case with only one thickness
        RAT = f(pol, n_list, d_list, th_0, lams, c_list)
        return RAT['R'], RAT['T']
        
    else:
        #case with varying thicknesses
        keys = thicks.keys()
        vals = thicks.values()
        sizes =  [len(val) for val in vals]
        params = product(*vals) #cartesian product of all thicknesses
        total = np.prod(sizes)
        # out = [np.nan]*total
        Rs = np.empty((total,n_lams))
        Ts = np.empty((total,n_lams))
        
        for i,par in enumerate(params):
            if not i%500: print(f'case {i}/{total}')
            
            for key,t in zip(keys, par):
                d_list[key] = t
            
            RAT = f(pol, n_list, d_list, th_0, lams, c_list)
            Rs[i] = RAT['R']
            Ts[i] = RAT['T']
            
        print('####finished####',)
        return Rs.reshape(*sizes,n_lams), Ts.reshape(*sizes,n_lams)
   

def RT_with_cache(filename, stack, materials, lams, pol='s', th_0=0, thicks=None):
    """Cache simple basado en el nombre de archivo. Los arreglos se guardan en el disco
    usando pickle."""
    if filename is None:
        return calculate_RT(stack, materials, lams, pol, th_0, thicks)
    else:
    
        if os.path.exists(filename):
            with open(filename,'rb') as f:
                print('loading from cache\n',)
                RT = pickle.load(f)
        else:
            RT = calculate_RT(stack, materials, lams, pol, th_0, thicks)
            with open(filename,'wb') as f:
                print('saving to cache\n')
                pickle.dump(RT, f)
        return RT

def calculate_ellips(n_list, d_list, lams, th_0):
    """parametros elipsometricos de elipsometro Horiba"""
        
    elipso_data = tmm.ellips(n_list, d_list, th_0/DEG, lams) 
        
    psi   = elipso_data["psi"]*DEG
    delta = -(elipso_data["Delta"]*DEG-180)

    return psi, delta


def ellips_tan_cos(n_list, d_list, th_0, lam_vac):
    """Tan y cos de los parámetros elipsométricos. Útil para ajustar.
    """

    s_data = tmm.coh_tmm('s', n_list, d_list, th_0/DEG, lam_vac)
    p_data = tmm.coh_tmm('p', n_list, d_list, th_0/DEG, lam_vac)
    rs = s_data['r']
    rp = p_data['r']
    div = rp/rs
    mod = abs(div)
    # return {'tanpsi': mod, 'cosDelta': np.real(div)/mod}
    return mod, np.real(div)/mod


def load_stack(filename, materials):
    """Carga un stack de archivos .xlsx usados por IR-S en Matlab. Los materiales NUEVOS 
    encontrados se agregaran al diccionario 'materials'. Cerciorarse que nk distintos
    tengan nombres distintos.
    """
    
    wb = load_workbook(filename = filename)
    ws = wb['MJSC Definition']
    
    stack = []
    start_row = 19
    end_row = 35
    
    for row in range(start_row, end_row+1):
        
        mat = ws.cell(row=row, column=3).value #C
        nk = ws.cell(row=row, column=5).value  #E
        thick = ws.cell(row=row, column=7).value #G
        
        if mat not in materials:
            # print('working on '+mat)
            materials[mat] = (load_interp('./nkdata/optical/' + nk))
        
        stack.append([thick, mat, 'c'])

    return  stack, materials
