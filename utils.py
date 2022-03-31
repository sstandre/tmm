# -*- coding: utf-8 -*-
"""
Created on Thu Mar 26 17:45:17 2020

@author: simon

Utilities to calculate RAT of optical stacks with tmm_vec
"""

import numpy as np
import os
import pickle
import tmm
from functools import partial
from itertools import product
from scipy.interpolate import interp1d

try:
    from openpyxl import load_workbook
    import matplotlib.pyplot as plt
    import matplotlib.cm as colors
except ImportError:
    pass


def cauchy_fn(A,B,C):
    def fn(lams):
        lams = lams*1.0
        return A+B*np.power(10,4)/np.power(lams,2)+C*np.power(10,9)/np.power(lams,4)
    return fn


def constant_fn(val):
    return partial(np.full_like, fill_value=val)


def brugg_fn(n_a, n_b, f_b):
    def fn(lams):
        n1 = n_a(lams)
        n2 = n_b(lams)
        return n_eff(n1, n2, f_b)
    return fn

def load_fn(filename, skiprows=1):
    
    wv, I = np.loadtxt(filename, skiprows=skiprows, unpack=True)
    return interp1d(wv, I,fill_value=(I[0],I[-1]), bounds_error=False)


def n_eff(n1,n2,f):
    """
    Modelo de bruggeman. f es la fraccion de la especie 2.
    """
    if f == 0: return n1
    if f == 1: return n2
    e1 = n1**2
    e2 = n2**2
    #f = 1-f2
    omega = (1-f)*(e2-2*e1)+f*(e1-2*e2)
    return (np.sqrt(np.sqrt(omega**2 + 8*e1*e2)-omega))/2


def load_interp(filename, comments=';', skiprows=0, unit='nm'):
    '''loads optical data and outputs interpolation functions for n and k. 
    Output functions take wv in nm. Keep default arguments to load .nkv files
    '''
    scale = {'nm':1.0, 'um':1000.0}
    
    wv, n, k = np.loadtxt(filename, comments=comments, 
                      skiprows=skiprows, unpack=True)
    
    wv *= scale[unit]
    
    n_fn = interp1d(wv, n,fill_value=(n[0],n[-1]), bounds_error=False)
    k_fn = interp1d(wv, k,fill_value=(k[0],k[-1]), bounds_error=False)
    
    return n_fn, k_fn


def load_psi_delta_from_spe(filename, offset):
    """Cargar datos elipsometricos de archivos .spe del elipsometro Horiba"""
    A = np.loadtxt(filename,skiprows=offset+62, usecols=(0,1,2), encoding="latin-1")
    L= A[:-2,0]
    P = A[:-2,:2]
    D = np.c_[L, A[:-2,2]]
    return P, D


def stack2tmm(stack, materials, lams, add_inf=True):
    '''Transforms a stack into three lists that can be fed into the 'tmm' package.
    BY DEFAULT ADDS inf AIR LAYERS'''
    n_list = []
    d_list = []
    c_list = []
    if add_inf:
        n_list.append([1.]*len(lams))
        d_list.append(np.inf)
        c_list.append('i')
        
    for layer in stack:
        d_list.append(layer[0])
        mat = materials[layer[1]]
        n_list.append(mat[0](lams) + mat[1](lams)*1.0j)
        c_list.append(layer[2])
        
    if add_inf:  
        n_list.append([1.]*len(lams))
        d_list.append(np.inf)
        c_list.append('i')
    
    return n_list, d_list, c_list

def calculate_RT(stack, materials, lams, pol='s', th_0=0, thicks=None):
    '''calulate R and T of a stack, for all combinations of thicks.'''
    n_list, d_list, c_list = stack2tmm(stack, materials, lams, add_inf=False)
    n_lams = len(lams)
    inc = 'i' in c_list[1:-1] #check if  any of the finite layers is incocherent
    if inc:
        def f(pol, n_list, d_list, th_0, lams, c_list):
            return tmm.inc_tmm(pol, n_list, d_list, c_list, th_0, lams)
    else:
        def f(pol, n_list, d_list, th_0, lams, c_list=None):
            return tmm.coh_tmm(pol, n_list, d_list, th_0, lams)
            
    if thicks is None:
        #case with only one thickness
        RAT = f(pol, n_list, d_list, th_0, lams, c_list)
        return RAT['R'], RAT['T']
        
    else:
        #case with varying thicknesses
        keys = thicks.keys()
        vals = thicks.values()
        sizes =  [len(val) for val in vals]
        params = product(*vals) #cartesian product of all thicknesses
        total = np.prod(sizes)
        # out = [np.nan]*total
        Rs = np.empty((total,n_lams))
        Ts = np.empty((total,n_lams))
        
        for i,par in enumerate(params):
            if not i%500: print(f'case {i}/{total}')
            
            for key,t in zip(keys, par):
                d_list[key] = t
            
            RAT = f(pol, n_list, d_list, th_0, lams, c_list)
            Rs[i] = RAT['R']
            Ts[i] = RAT['T']
            
        print('####finished####',)
        return Rs.reshape(*sizes,n_lams), Ts.reshape(*sizes,n_lams)
   

def RT_with_cache(filename, stack, materials, lams, pol='s', th_0=0, thicks=None):
    """Simple filename-based cache for RT data. Arrays are pickled to disk."""
    if filename is None:
        return calculate_RT(stack, materials, lams, pol, th_0, thicks)
    else:
    
        if os.path.exists(filename):
            with open(filename,'rb') as f:
                print('loading from cache\n',)
                RT = pickle.load(f)
        else:
            RT = calculate_RT(stack, materials, lams, pol, th_0, thicks)
            with open(filename,'wb') as f:
                print('saving to cache\n')
                pickle.dump(RT, f)
        return RT

def calculate_ellips(n_list, d_list, lams, th_0):
    """parametros elipsometricos de elipsometro Horiba"""
        
    elipso_data = tmm.ellips(n_list, d_list, th_0*np.pi/180, lams) 
        
    psi   = elipso_data["psi"]*180/np.pi
    delta = -(elipso_data["Delta"]*180/np.pi-180)

    return psi, delta



def load_stack(filename, materials):
    '''Loads stack from Xlsx used by IR-S in Matlab. 'Materials' is a dict 
    where new materials will be added. Only materials not already present will 
    be added. YOU SHOULD CHECK that different nk's have different material names
    '''
    
    wb = load_workbook(filename = filename)
    ws = wb['MJSC Definition']
    
    stack = []
    start_row = 19
    end_row = 35
    
    for row in range(start_row, end_row+1):
        
        mat = ws.cell(row=row, column=3).value #C
        nk = ws.cell(row=row, column=5).value  #E
        thick = ws.cell(row=row, column=7).value #G
        
        if mat not in materials:
            # print('working on '+mat)
            materials[mat] = (load_interp('./nkdata/optical/' + nk))
        
        stack.append([thick, mat, 'c'])

    return  stack, materials
